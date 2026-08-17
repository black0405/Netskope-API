"""
Migrate old hand-named GenAI Analysis traffic files to the format produced
by export_genai_page_bytes.py.

    python migrate_old_traffic.py [--dry-run]

Run it from inside the folder containing the old files.

Handles filenames like:
    GenAI Analysis 17th Aug 2026 053716UTC.csv
    GenAI Analysis 14-18feb.csv
    GenAI_Analysis_2025-04-30T0031.csv
    GenAI Analysis 29may.csv

Each becomes GenerativeAI_Traffic_YYYY-MM-DD.csv (first day of a range;
year from the name when present, else 2026), and the columns are reshaped
to the export layout:
    Application, User, URL, Event Date, Organization Unit,
    Bytes Upload, Bytes Download, Total Bytes
Extra columns (User Group, Department...) are dropped. A file is skipped
only if one of the export columns is missing entirely.

.xlsx files are handled too (first worksheet); their sheet is renamed to
EXCEL_SHEET_NAME to match the export script's output. CSV output stays
CSV -- the format has no sheet name.
"""

import csv
import re
import sys
from pathlib import Path

DEFAULT_YEAR = 2026

# Some old files hold huge fields (long URLs / embedded blobs); the default
# 128KB csv limit chokes on them. 1e9 stays within Windows' 32-bit C long.
csv.field_size_limit(1_000_000_000)

# Same worksheet name export_genai_page_bytes.py uses for .xlsx output.
EXCEL_SHEET_NAME = "in"

TARGET_HEADERS = ["Application", "User", "URL", "Event Date",
                  "Organization Unit", "Bytes Upload", "Bytes Download",
                  "Total Bytes"]

MONTHS = {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
          "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12}


def parse_date(name: str) -> str:
    """Pull a YYYY-MM-DD label out of an old filename, or raise ValueError."""
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", name)
    if m:
        return m.group(0)
    # "29may" / "14-18feb" / "17th Aug 2026"
    # -> first day + month word + optional year
    m = re.search(r"(\d{1,2})(?:st|nd|rd|th)?(?:\s*-\s*\d{1,2})?[\s_-]*"
                  r"([a-zA-Z]+)(?:[\s_-]+(\d{4}))?", name)
    if m:
        month = MONTHS.get(m.group(2).lower()[:3])
        if month:
            year = int(m.group(3)) if m.group(3) else DEFAULT_YEAR
            return f"{year}-{month:02d}-{int(m.group(1)):02d}"
    raise ValueError(f"no date found in {name!r}")


def fix_header(old: list) -> list:
    """Map old header variants onto the canonical export headers."""
    canon = {re.sub(r"[^a-z]", "", h.lower()): h for h in TARGET_HEADERS}
    canon[re.sub(r"[^a-z]", "", "Event Data".lower())] = "Event Date"
    return [canon.get(re.sub(r"[^a-z]", "", h.lower()), h) for h in old]


def read_rows(path: Path) -> list:
    if path.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook
        ws = load_workbook(path).worksheets[0]
        return [["" if c is None else c for c in row]
                for row in ws.iter_rows(values_only=True)]
    # Old files come from mixed sources: try UTF-8 first, fall back to
    # cp1252 (Excel "ANSI" saves), which accepts any byte sequence.
    for enc in ("utf-8-sig", "cp1252"):
        try:
            with open(path, newline="", encoding=enc) as f:
                return list(csv.reader(f))
        except UnicodeDecodeError:
            continue


def write_rows(path: Path, rows: list) -> None:
    if path.suffix.lower() == ".xlsx":
        from openpyxl import Workbook
        wb = Workbook()
        wb.active.title = EXCEL_SHEET_NAME
        for row in rows:
            wb.active.append(row)
        wb.save(path)
        return
    with open(path, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(rows)


def migrate(folder: Path, dry_run: bool) -> None:
    old_files = [p for ext in ("*.csv", "*.xlsx") for p in folder.glob(ext)
                 if re.match(r"GenAI[ _]Analysis", p.name, re.I)]
    if not old_files:
        print(f"No 'GenAI Analysis' files found in {folder}")
        return
    for path in sorted(old_files):
        try:
            label = parse_date(path.stem)
        except ValueError as exc:
            print(f"SKIP {path.name}: {exc}")
            continue
        rows = read_rows(path)
        fixed = fix_header([str(h) for h in rows[0]]) if rows else []
        missing = [h for h in TARGET_HEADERS if h not in fixed]
        if missing:
            print(f"SKIP {path.name}: missing columns {missing}")
            continue
        ext = path.suffix.lower()
        dest = folder / f"GenerativeAI_Traffic_{label}{ext}"
        n = 2
        while dest.exists():                     # never clobber
            dest = folder / f"GenerativeAI_Traffic_{label}_{n}{ext}"
            n += 1
        print(f"{path.name}  ->  {dest.name}")
        if dry_run:
            continue
        # Reshape every row to the export layout: the eight columns in
        # order; anything else (User Group, Department...) drops.
        idx = [fixed.index(h) for h in TARGET_HEADERS]
        out = [TARGET_HEADERS] + [
            [row[j] if j < len(row) else "" for j in idx]
            for row in rows[1:] if row
        ]
        write_rows(dest, out)
        path.unlink()


def selftest() -> None:
    assert parse_date("GenAI Analysis 17th Aug 2026 053716UTC") == "2026-08-17"
    assert parse_date("GenAI Analysis 14-18feb") == "2026-02-14"
    assert parse_date("GenAI_Analysis_2025-04-30T0031") == "2025-04-30"
    assert parse_date("GenAI Analysis 29may") == "2026-05-29"
    assert fix_header(["Application", "Event Data", "Bytes Upload"]) == \
        ["Application", "Event Date", "Bytes Upload"]
    print("selftest OK")


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if a != "--dry-run"]
    if "--selftest" in args:
        selftest()
        sys.exit(0)
    migrate(Path(args[0]) if args else Path.cwd(), dry_run="--dry-run" in sys.argv)
