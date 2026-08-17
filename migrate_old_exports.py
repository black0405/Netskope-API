"""
Migrate old hand-named GenAI export files to the formats produced by
export_genai_applications.py and export_genai_page_bytes.py.

    python migrate_old_exports.py [--dry-run]

Run it from inside the folder containing the old files.

Two file families are recognised by name:

  "GENAI apps activity ..."  -> GenerativeAI_Applications_YYYY-MM-DD
      S.No., User, Application, URL, Activity, Object Type, Object Name,
      Event Date, Organization Unit, Sum - File Size (MB)

  "GenAI Analysis ..."       -> GenerativeAI_Traffic_YYYY-MM-DD
      Application, User, URL, Event Date, Organization Unit,
      Bytes Upload, Bytes Download, Total Bytes    (no S.No. column)

Date is pulled from the filename: ISO (2026-03-05), "10 march", "14-19jan"
(first day of a range), "17th Aug 2026", "29may". Year defaults to 2026
when the name has none.

Columns are reshaped to the export layout: header variants fixed
(Event Data -> Event Date, Sum - File Size(MB) -> Sum - File Size (MB)),
extra columns (department, User Group...) dropped, S.No. renumbered 1..n
where the family has one. A file is skipped only if one of its family's
columns is missing entirely.

.xlsx files are handled too (first worksheet); their sheet is renamed to
EXCEL_SHEET_NAME to match the export scripts. CSV output stays CSV --
the format has no sheet name.
"""

import csv
import re
import sys
from pathlib import Path

DEFAULT_YEAR = 2026

# Some old files hold huge fields (long URLs / embedded blobs); the default
# 128KB csv limit chokes on them. 1e9 stays within Windows' 32-bit C long.
csv.field_size_limit(1_000_000_000)

# Same worksheet name the export scripts use for .xlsx output.
EXCEL_SHEET_NAME = "in"

# Each family: filename pattern, output stem, exact target header row.
# A leading "S.No." means the family carries a generated row counter.
FAMILIES = [
    {
        "match": r"GENAI[ _]apps[ _]activity",
        "out": "GenerativeAI_Applications",
        "headers": ["S.No.", "User", "Application", "URL", "Activity",
                    "Object Type", "Object Name", "Event Date",
                    "Organization Unit", "Sum - File Size (MB)"],
    },
    {
        "match": r"GenAI[ _]Analysis",
        "out": "GenerativeAI_Traffic",
        "headers": ["Application", "User", "URL", "Event Date",
                    "Organization Unit", "Bytes Upload", "Bytes Download",
                    "Total Bytes"],
    },
]

MONTHS = {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
          "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12}


def parse_date(name: str) -> str:
    """Pull a YYYY-MM-DD label out of an old filename, or raise ValueError."""
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", name)
    if m:
        return m.group(0)
    # "10 march" / "14-19jan" / "17th Aug 2026" / "29may"
    # -> first day + month word + optional year
    m = re.search(r"(\d{1,2})(?:st|nd|rd|th)?(?:\s*-\s*\d{1,2})?[\s_-]*"
                  r"([a-zA-Z]+)(?:[\s_-]+(\d{4}))?", name)
    if m:
        month = MONTHS.get(m.group(2).lower()[:3])
        if month:
            year = int(m.group(3)) if m.group(3) else DEFAULT_YEAR
            return f"{year}-{month:02d}-{int(m.group(1)):02d}"
    raise ValueError(f"no date found in {name!r}")


# Every known header, keyed by its letters-only lowercase form, so spacing
# and punctuation variants all land on the canonical name.
_ALL_HEADERS = [h for fam in FAMILIES for h in fam["headers"]]
_CANON = {re.sub(r"[^a-z]", "", h.lower()): h for h in _ALL_HEADERS}
_CANON[re.sub(r"[^a-z]", "", "Event Data".lower())] = "Event Date"


def fix_header(old: list) -> list:
    """Map old header variants onto the canonical export headers."""
    out = []
    for i, h in enumerate(old):
        if i == 0 and not h.strip():
            out.append("S.No.")
        else:
            out.append(_CANON.get(re.sub(r"[^a-z]", "", h.lower()), h))
    return out


def read_rows(path: Path) -> list:
    if path.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook
        ws = load_workbook(path).worksheets[0]
        return [["" if c is None else c for c in row]
                for row in ws.iter_rows(values_only=True)]
    # Old files come from mixed sources: try UTF-8 first, fall back to
    # cp1252 (Excel "ANSI" saves), which accepts any byte sequence.
    for enc in ("utf-8-sig", "cp1252"):
        try:
            with open(path, newline="", encoding=enc) as f:
                return list(csv.reader(f))
        except UnicodeDecodeError:
            continue


def write_rows(path: Path, rows: list) -> None:
    if path.suffix.lower() == ".xlsx":
        from openpyxl import Workbook
        wb = Workbook()
        wb.active.title = EXCEL_SHEET_NAME
        for row in rows:
            wb.active.append(row)
        wb.save(path)
        return
    with open(path, "w", newline="", encoding="utf-8") as f:
        csv.writer(f).writerows(rows)


def family_for(name: str):
    for fam in FAMILIES:
        if re.match(fam["match"], name, re.I):
            return fam
    return None


def migrate(folder: Path, dry_run: bool) -> None:
    old_files = [p for ext in ("*.csv", "*.xlsx") for p in folder.glob(ext)
                 if family_for(p.name)]
    if not old_files:
        print(f"No old GenAI export files found in {folder}")
        return
    for path in sorted(old_files):
        fam = family_for(path.name)
        try:
            label = parse_date(path.stem)
        except ValueError as exc:
            print(f"SKIP {path.name}: {exc}")
            continue
        rows = read_rows(path)
        fixed = fix_header([str(h) for h in rows[0]]) if rows else []
        serial = fam["headers"][0] == "S.No."
        data_cols = fam["headers"][1:] if serial else fam["headers"]
        missing = [h for h in data_cols if h not in fixed]
        if missing:
            print(f"SKIP {path.name}: missing columns {missing}")
            continue
        ext = path.suffix.lower()
        dest = folder / f"{fam['out']}_{label}{ext}"
        n = 2
        while dest.exists():                     # never clobber
            dest = folder / f"{fam['out']}_{label}_{n}{ext}"
            n += 1
        print(f"{path.name}  ->  {dest.name}")
        if dry_run:
            continue
        # Reshape every row to the export layout: the family's columns in
        # order (plus a fresh S.No. counter where it has one); anything
        # else (department, User Group...) drops.
        idx = [fixed.index(h) for h in data_cols]
        out = [fam["headers"]] + [
            ([i] if serial else [])
            + [row[j] if j < len(row) else "" for j in idx]
            for i, row in enumerate(rows[1:], 1) if row
        ]
        write_rows(dest, out)
        path.unlink()


def selftest() -> None:
    assert parse_date("GENAI apps activity 10 march") == "2026-03-10"
    assert parse_date("GENAI apps activity 14-19jan") == "2026-01-14"
    assert parse_date("GENAI_apps_activity_2026-03-05T0601") == "2026-03-05"
    assert parse_date("GenAI Analysis 17th Aug 2026 053716UTC") == "2026-08-17"
    assert parse_date("GenAI Analysis 14-18feb") == "2026-02-14"
    assert parse_date("GenAI_Analysis_2025-04-30T0031") == "2025-04-30"
    assert parse_date("GenAI Analysis 29may") == "2026-05-29"
    assert fix_header(["", "User", "Event Data", "Sum - File Size(MB)"]) == \
        ["S.No.", "User", "Event Date", "Sum - File Size (MB)"]
    assert family_for("GenAI Analysis 29may.csv")["out"] == \
        "GenerativeAI_Traffic"
    assert family_for("GENAI apps activity 10 march.csv")["out"] == \
        "GenerativeAI_Applications"
    assert family_for("random.csv") is None
    print("selftest OK")


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if a != "--dry-run"]
    if "--selftest" in args:
        selftest()
        sys.exit(0)
    migrate(Path(args[0]) if args else Path.cwd(), dry_run="--dry-run" in sys.argv)
